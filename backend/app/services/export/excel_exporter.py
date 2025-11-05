"""
Excel 导出服务
支持自定义字段选择、预设模板、临时文件管理
"""
import os
import tempfile
import json
from datetime import datetime
from typing import List, Dict, Any, Optional
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.core.utils.logger import app_logger as logger


class ExcelExporter:
    """Excel导出器"""
    
    # 预设模板配置
    TEMPLATES = {
        'basic': {
            'name': '基础信息',
            'fields': ['ip.ip_address', 'hostname', 'host_type', 'os_name', 'collection_status']
        },
        'detailed': {
            'name': '详细信息',
            'fields': [
                'ip.ip_address', 'hostname', 'host_type', 'os_name', 'os_version', 
                'cpu_model', 'cpu_cores', 'memory_total', 'collection_status'
            ]
        },
        'full': {
            'name': '完整信息',
            'fields': [
                'ip.ip_address', 'hostname', 'host_type', 'os_name', 'os_version', 
                'kernel_version', 'cpu_model', 'cpu_cores', 'memory_total',
                'network_interfaces', 'disk_info', 'vmware_info', 'collection_status',
                'last_collected_at', 'collection_error'
            ]
        }
    }
    
    def __init__(self):
        self.export_dir = None
        self.file_expiry = 3600  # 默认1小时过期
    
    def init_app(self, app):
        """初始化应用配置"""
        with app.app_context():
            self.export_dir = app.config.get('EXPORT_FILE_DIR', tempfile.gettempdir())
            self.file_expiry = app.config.get('EXPORT_FILE_EXPIRY', 3600)
            
            # 确保导出目录存在
            os.makedirs(self.export_dir, exist_ok=True)
    
    def export_hosts(self, hosts: List[Dict[str, Any]], fields: List[str], 
                    template: Optional[str] = None) -> str:
        """
        导出主机信息到Excel
        支持导出父主机及其子主机，自动添加父主机信息列
        
        Args:
            hosts: 主机信息列表（包含父主机和子主机）
            fields: 要导出的字段列表
            template: 预设模板名称（可选）
            
        Returns:
            导出文件的路径
        """
        try:
            # 如果有预设模板，使用模板字段
            if template and template in self.TEMPLATES:
                fields = self.TEMPLATES[template]['fields']
            
            # 检测是否有子主机（通过检查是否有_parent_hostname字段）
            has_children = any('_parent_hostname' in host for host in hosts)
            
            # 如果有子主机，在字段列表前添加"父主机"列
            export_fields = fields.copy()
            parent_hostname_index = 0
            if has_children:
                # 将父主机信息列添加到最前面
                export_fields.insert(0, '_parent_hostname')
                parent_hostname_index = 0
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "主机信息"
            
            # 设置表头
            headers = self._get_field_headers(export_fields)
            ws.append(headers)
            
            # 设置表头样式
            self._style_headers(ws, headers)
            
            # 按父主机和子主机分组排序：先父主机，后子主机
            sorted_hosts = self._sort_hosts_by_hierarchy(hosts)
            
            # 填充数据
            for host in sorted_hosts:
                row_data = self._extract_row_data(host, export_fields)
                ws.append(row_data)
            
            # 调整列宽
            self._adjust_column_width(ws, headers)
            
            # 保存文件
            filename = f"hosts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            wb.save(filepath)
            
            logger.info(f"Excel export completed: {filepath}, total hosts: {len(hosts)}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def _get_field_headers(self, fields: List[str]) -> List[str]:
        """
        获取字段的显示名称
        
        Args:
            fields: 字段列表
            
        Returns:
            表头列表
        """
        field_names = {
            '_parent_hostname': '父主机',  # 新增：父主机信息列
            'ip.ip_address': 'IP地址',
            'hostname': '主机名',
            'host_type': '主机类型',
            'os_name': '操作系统',
            'os_version': '系统版本',
            'kernel_version': '内核版本',
            'cpu_model': 'CPU型号',
            'cpu_cores': 'CPU核心数',
            'memory_total': '总内存(MB)',
            'network_interfaces': '网络接口',
            'disk_info': '磁盘信息',
            'vmware_info': 'VMware信息',
            'collection_status': '采集状态',
            'last_collected_at': '最后采集时间',
            'collection_error': '采集错误'
        }
        
        return [field_names.get(field, field) for field in fields]
    
    def _extract_row_data(self, host: Dict[str, Any], fields: List[str]) -> List[Any]:
        """
        从主机数据中提取行数据
        
        Args:
            host: 主机数据字典
            fields: 字段列表
            
        Returns:
            行数据列表
        """
        row_data = []
        
        for field in fields:
            value = None
            
            # 处理父主机信息字段
            if field == '_parent_hostname':
                value = host.get('_parent_hostname', '')
            # 处理主机名字段：对于VMware，优先显示vm_name，否则显示hostname
            elif field == 'hostname':
                # 如果是VMware类型，从vmware_info中获取vm_name
                if host.get('host_type') == 'vmware' and host.get('vmware_info'):
                    vmware_info = host.get('vmware_info', {})
                    if isinstance(vmware_info, dict):
                        value = vmware_info.get('vm_name') or host.get('hostname')
                    else:
                        value = host.get('hostname')
                else:
                    value = host.get('hostname')
            # 从host_info或ip中获取值
            elif '.' in field:
                # 嵌套字段，如 ip.ip_address
                parts = field.split('.')
                value = host
                for part in parts:
                    if isinstance(value, dict):
                        value = value.get(part)
                    else:
                        value = None
                        break
            else:
                value = host.get(field)
            
            # 处理特殊字段
            if field == 'memory_total' and value:
                value = f"{value} MB"
            elif field == 'disk_info' and value:
                # 使用格式化函数格式化磁盘信息
                host_type = host.get('host_type', 'physical')
                value = self._format_disk_info(value, host_type)
            elif field == 'network_interfaces' and value:
                # 使用格式化函数格式化网络信息
                host_type = host.get('host_type', 'physical')
                value = self._format_network_info(value, host_type)
            elif field == 'vmware_info' and value:
                # JSON字段转换为字符串
                value = json.dumps(value, ensure_ascii=False)
            elif field == 'collection_status' and value:
                # 状态转换
                status_map = {
                    'pending': '待采集',
                    'collecting': '采集中',
                    'success': '成功',
                    'failed': '失败'
                }
                value = status_map.get(value, value)
            elif field == 'host_type' and value:
                # 主机类型转换
                type_map = {
                    'physical': '物理机',
                    'vmware': 'VMware',
                    'other_virtualization': '其他虚拟化'
                }
                value = type_map.get(value, value)
            elif field == 'last_collected_at' and value:
                # 日期格式化
                if isinstance(value, str):
                    value = value
                else:
                    value = value.isoformat() if hasattr(value, 'isoformat') else str(value)
            
            row_data.append(value)
        
        return row_data
    
    def _sort_hosts_by_hierarchy(self, hosts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        按层级关系排序主机：先父主机，后子主机
        
        Args:
            hosts: 主机列表
            
        Returns:
            排序后的主机列表
        """
        # 分离父主机和子主机
        parent_hosts = []
        child_hosts = []
        
        for host in hosts:
            if host.get('parent_host_id') or host.get('_parent_hostname'):
                child_hosts.append(host)
            else:
                parent_hosts.append(host)
        
        # 构建父主机ID到子主机的映射
        parent_id_map = {}
        for child in child_hosts:
            parent_id = child.get('parent_host_id')
            if parent_id:
                if parent_id not in parent_id_map:
                    parent_id_map[parent_id] = []
                parent_id_map[parent_id].append(child)
        
        # 排序：先父主机，然后按父主机ID分组添加子主机
        sorted_hosts = []
        for parent in parent_hosts:
            sorted_hosts.append(parent)
            # 添加该父主机的所有子主机
            parent_id = parent.get('id')
            if parent_id in parent_id_map:
                sorted_hosts.extend(parent_id_map[parent_id])
        
        # 如果有孤立子主机（父主机不在列表中），也添加它们
        processed_child_ids = set()
        for parent_id in parent_id_map:
            for child in parent_id_map[parent_id]:
                processed_child_ids.add(child.get('id'))
        
        for child in child_hosts:
            if child.get('id') not in processed_child_ids:
                sorted_hosts.append(child)
        
        return sorted_hosts
    
    def _style_headers(self, ws, headers: List[str]):
        """
        设置表头样式
        
        Args:
            ws: 工作表对象
            headers: 表头列表
        """
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _adjust_column_width(self, ws, headers: List[str]):
        """
        调整列宽
        
        Args:
            ws: 工作表对象
            headers: 表头列表
        """
        for col in range(1, len(headers) + 1):
            # 获取最大内容长度
            max_length = len(str(headers[col - 1]))
            
            # 检查当前列的所有单元格
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # 设置列宽
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    def get_available_fields(self) -> List[Dict[str, str]]:
        """
        获取所有可用的导出字段
        
        Returns:
            字段列表，包含字段名和显示名称
        """
        return [
            {'field': 'ip.ip_address', 'label': 'IP地址', 'category': 'basic'},
            {'field': 'hostname', 'label': '主机名', 'category': 'basic'},
            {'field': 'host_type', 'label': '主机类型', 'category': 'basic'},
            {'field': 'os_name', 'label': '操作系统', 'category': 'system'},
            {'field': 'os_version', 'label': '系统版本', 'category': 'system'},
            {'field': 'kernel_version', 'label': '内核版本', 'category': 'system'},
            {'field': 'cpu_model', 'label': 'CPU型号', 'category': 'hardware'},
            {'field': 'cpu_cores', 'label': 'CPU核心数', 'category': 'hardware'},
            {'field': 'memory_total', 'label': '总内存(MB)', 'category': 'hardware'},
            {'field': 'network_interfaces', 'label': '网络接口', 'category': 'network'},
            {'field': 'disk_info', 'label': '磁盘信息', 'category': 'storage'},
            {'field': 'vmware_info', 'label': 'VMware信息', 'category': 'vmware'},
            {'field': 'collection_status', 'label': '采集状态', 'category': 'status'},
            {'field': 'last_collected_at', 'label': '最后采集时间', 'category': 'status'},
            {'field': 'collection_error', 'label': '采集错误', 'category': 'status'}
        ]
    
    def _format_disk_info(self, disk_info: Any, host_type: str = 'physical') -> str:
        """
        格式化磁盘信息为管道分隔字符串
        
        Args:
            disk_info: 磁盘信息（可能是列表或JSON字符串）
            host_type: 主机类型（vmware, physical等）
            
        Returns:
            格式化后的字符串，每行一个磁盘，字段用|分隔
        """
        try:
            # 如果是字符串，尝试解析JSON
            if isinstance(disk_info, str):
                disk_info = json.loads(disk_info)
            
            if not disk_info or not isinstance(disk_info, list):
                return ""
            
            formatted_lines = []
            
            for disk in disk_info:
                if not isinstance(disk, dict):
                    continue
                
                if host_type == 'vmware':
                    # VMware格式：datastore|file_name|capacity_kb|disk_mode
                    datastore = disk.get('datastore', '')
                    file_name = disk.get('file_name', '')
                    capacity_kb = disk.get('capacity_kb', 0)
                    disk_mode = disk.get('disk_mode', '')
                    
                    # 转换为GB
                    capacity_gb = f"{capacity_kb / 1024 / 1024:.2f}GB" if capacity_kb else "0GB"
                    
                    line = f"{datastore}|{file_name}|{capacity_gb}|{disk_mode}"
                    formatted_lines.append(line)
                else:
                    # Linux/Windows格式：device|size|vendor|model|mount|fstype|use_percent
                    device = disk.get('device') or disk.get('DeviceID', '')
                    size = disk.get('size') or disk.get('Size', '')
                    
                    # Linux格式处理
                    if 'vendor' in disk:
                        vendor = disk.get('vendor', '')
                        model = disk.get('model', '')
                        mount = disk.get('mount', '')
                        fstype = disk.get('fstype', '')
                        use_percent = disk.get('use_percent', '')
                        line = f"{device}|{size}|{vendor}|{model}|{mount}|{fstype}|{use_percent}"
                    # Windows格式处理
                    elif 'Model' in disk:
                        model = disk.get('Model', '')
                        interface = disk.get('InterfaceType', '')
                        media_type = disk.get('MediaType', '')
                        line = f"{device}|{size}GB|{model}|{interface}|{media_type}"
                    else:
                        # 通用格式：device|size|free_space|used_space|use_percent
                        free_space = disk.get('FreeSpace') or disk.get('avail', '')
                        used_space = disk.get('UsedSpace') or disk.get('used', '')
                        use_percent = disk.get('UsePercent') or disk.get('use_percent', '')
                        line = f"{device}|{size}|{free_space}|{used_space}|{use_percent}%"
                    
                    formatted_lines.append(line)
            
            return '\n'.join(formatted_lines) if formatted_lines else ""
            
        except Exception as e:
            logger.warning(f"Error formatting disk info: {str(e)}")
            # 如果格式化失败，返回JSON字符串
            if isinstance(disk_info, str):
                return disk_info
            return json.dumps(disk_info, ensure_ascii=False) if disk_info else ""
    
    def _format_network_info(self, network_interfaces: Any, host_type: str = 'physical') -> str:
        """
        格式化网络信息为管道分隔字符串
        
        Args:
            network_interfaces: 网络接口信息（可能是列表或JSON字符串）
            host_type: 主机类型（vmware, physical等）
            
        Returns:
            格式化后的字符串，每行一个网卡，字段用|分隔
        """
        try:
            # 如果是字符串，尝试解析JSON
            if isinstance(network_interfaces, str):
                network_interfaces = json.loads(network_interfaces)
            
            if not network_interfaces or not isinstance(network_interfaces, list):
                return ""
            
            formatted_lines = []
            
            for interface in network_interfaces:
                if not isinstance(interface, dict):
                    continue
                
                if host_type == 'vmware':
                    # VMware格式：network_name|mac_address|connected|ip_addresses
                    network_name = interface.get('network_name', '')
                    mac_address = interface.get('mac_address', '')
                    connected = interface.get('connected', '')
                    ip_addresses = interface.get('ip_addresses', [])
                    
                    # IP地址列表转换为字符串
                    ip_str = ','.join(ip_addresses) if isinstance(ip_addresses, list) else str(ip_addresses)
                    
                    line = f"{network_name}|{mac_address}|{connected}|{ip_str}"
                    formatted_lines.append(line)
                else:
                    # Linux/Windows格式：name|mac_address|state|mtu|speed|ipv4|gateway
                    name = interface.get('name') or interface.get('Name', '')
                    mac_address = interface.get('mac_address') or interface.get('MacAddress', '')
                    
                    # Linux格式
                    if 'state' in interface:
                        state = interface.get('state', '')
                        mtu = interface.get('mtu', '')
                        speed = interface.get('speed', '')
                        ipv4 = interface.get('ipv4', '')
                        gateway = interface.get('gateway', '')
                        line = f"{name}|{mac_address}|{state}|{mtu}|{speed}|{ipv4}|{gateway}"
                    # Windows格式
                    elif 'Status' in interface:
                        status = interface.get('Status', '')
                        link_speed = interface.get('LinkSpeed', '')
                        ipv4 = interface.get('IPv4Address', '')
                        gateway = interface.get('Gateway', '')
                        line = f"{name}|{mac_address}|{status}|{link_speed}|{ipv4}|{gateway}"
                    else:
                        # 通用格式
                        line = f"{name}|{mac_address}"
                    
                    formatted_lines.append(line)
            
            return '\n'.join(formatted_lines) if formatted_lines else ""
            
        except Exception as e:
            logger.warning(f"Error formatting network info: {str(e)}")
            # 如果格式化失败，返回JSON字符串
            if isinstance(network_interfaces, str):
                return network_interfaces
            return json.dumps(network_interfaces, ensure_ascii=False) if network_interfaces else ""
    
    def get_template_list(self) -> List[Dict[str, Any]]:
        """
        获取预设模板列表
        
        Returns:
            模板列表，包含id、name、field_count和fields
        """
        return [
            {
                'id': key,
                'name': value['name'],
                'field_count': len(value['fields']),
                'fields': value['fields']  # 包含字段列表
            }
            for key, value in self.TEMPLATES.items()
        ]


# 创建全局实例
excel_exporter = ExcelExporter()

